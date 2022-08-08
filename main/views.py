from django.http import HttpResponse
from django.shortcuts import render,redirect
from django.conf import settings
from . import models
import json
import openpyxl
from django.views.decorators.csrf import csrf_exempt
from main.models import product,Invoice,Customer
curl = settings.CURRENT_URL
media_url=settings.MEDIA_URL
import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os
from openpyxl import Workbook
date = datetime.datetime.now().date()
today = date.today()
d2 = str(today.strftime("%d-%m-%Y"))

def Home(request):
    now = datetime.datetime.now()
    out=product.objects.all().order_by('pname')
    out1=Customer.objects.all().order_by('name')
    return render(request, "index0.html", {'curl': curl,'out':out,'out1':out1})


#///////////////////////////////////////////////////////
def viewfile(request):
    return render(request, "billview.html", {'curl': curl})



#///////////////////////////////////////////////////////
def Product(request):
    if request.method == "POST":
        hsncode = request.POST.get('hsn', '')
        pname = request.POST.get('pname', '')
        MRP = request.POST.get('mrp','')
        rate = request.POST.get('rate','')
        cgst = request.POST.get('cgst','')
        sgst = request.POST.get('sgst','')
        pro = product(hsncode=hsncode, pname=pname, MRP=MRP, rate=rate, cgst=cgst, sgst=sgst)
        pro.save()
        return render(request, "AddProduct.html", {'curl': curl, 'output': 'Added successfully'})
    else:
        return render(request, "AddProduct.html", {'curl': curl, 'output': ''})



#///////////////////////////////////////////////////////

def customer(request):
    if request.method == "POST":
        #id = request.POST.get('id')
        name = request.POST.get('name', '')
        add = request.POST.get('add','')
        pno = request.POST.get('pno','NULL')
        gst = request.POST.get('gst','')
        cust = Customer(name=name,add=add,pno=int(pno),gst=gst)
        cust.save()
        return render(request, "AddCustomer.html", {'curl': curl, 'output': 'Added successfully'})
    else:
        return render(request, "AddCustomer.html", {'curl': curl, 'output': ''})


#///////////////////////////////////////////////////////
def showcustomer(request):
    a=Customer.objects.values('id','name','add','pno','gst')
    return render(request, "showcustomer.html", {'curl': curl, 'out': a})


#///////////////////////////////////////////////////////
def showproduct(request):
    #a=Customer.objects.values('id','name','add','pno','gst')
    out=product.objects.values('id','hsncode','pname','MRP','rate','cgst','sgst')
    count=len(out)
    return render(request, "showproduct.html", {'curl': curl, 'out': out,'count':count})


#///////////////////////////////////////////////////////
def manageP(request):
    id = request.GET.get('e')
    name = request.GET.get('rid')
    action = request.GET.get('action')
    if action == 'delete':
        product.objects.get(pname=name,id=id).delete()
    return redirect(curl + "showproduct/")

#///////////////////////////////////////////////////////
def manageC(request):
    id = request.GET.get('e')
    name = request.GET.get('rid')
    action = request.GET.get('action')
    if action == 'delete':
        Customer.objects.get(name=name,id=id).delete()
    return redirect(curl + "showcustomer/")

#///////////////////////////////////////////////////////
@csrf_exempt
def checker(request):

    if request.method == "POST":
        invoice_no=Invoice.objects.values('id').last()
        invoice_no=invoice_no['id']
        key = json.loads(request.POST['key1'])
        key2 = json.loads(request.POST['key2'])
        cu_name=str(key2['cn'])
        cu_add=str(key2['ca'])
        cu_pno=str(key2['cm'])
        cu_gst=str(key2['cg'])
        flag_print=int(key2['a'])
        product_code=[]
        products_list=[]
        product_mrp=[]
        product_quantity=[]
        product_rate=[]
        product_dis=[]
        product_price=[]
        product_cgst=[]
        product_cgstp=[]
        product_sgst=[]
        product_sgstp=[]
        product_amount=[]
        akm=[]
        print("printing bill number "+str(invoice_no)+" of "+str(cu_name))
        for i in range(0,len(key)):
            akm.append(float(key[i]['akm']))
            product_code.append(key[i]['hsnc'])
            products_list.append(key[i]['pname'])
            product_mrp.append(float(key[i]['mrp']))
            product_quantity.append(int(key[i]['quan']))
            product_rate.append(float(key[i]['rate']))
            product_dis.append(int(key[i]['dis']))
            product_price.append(float(key[i]['ta']))
            product_cgst.append(int(key[i]['cgst']))
            product_cgstp.append(float(key[i]['cgstv']))
            product_sgst.append(int(key[i]['sgst']))
            product_sgstp.append(float(key[i]['sgstv']))
            product_amount.append(float(key[i]['amount']))
        tyiw = float(sum(akm))
        wb = Workbook()
        sheet = wb.active
        fontStyle = Font(name='Tahoma',size=10)
        fontScyil = Font(name='Comic Sans MS', size=9)
        fontamount = Font(name='Tahoma',size=12)
        ali = Alignment(horizontal='center',shrink_to_fit=True,indent=0)
        alia = Alignment(shrink_to_fit=True,indent=0)
        fontS = Font(name='Comic Sans MS',size=9)
        alie = Alignment(horizontal='center',shrink_to_fit=True,indent=0)
        now = datetime.datetime.now()
        hr=now.hour
        mn=now.minute
        if(str(mn)=='0' or str(mn)=='1'or str(mn)=='2'or str(mn)=='3'or str(mn)=='4'or str(mn)=='5'or str(mn)=='6'or str(mn)=='7'or str(mn)=='8'or str(mn)=='9'):
            mn='0'+str(mn)
        sc=now.second
        today = date.today()
        d1 = str(today.strftime("%d-%b-%Y"))
        file_name2 = str(today.strftime("%b-%y"))
        direct1 = "D:/Store Management web/MonthlyReport/" + str(file_name2) + "/"
        directory = "D:/Store Management web/Invoice/" + str(d1) + "/"
        if not os.path.exists(directory):
            os.makedirs(directory)
        bill_number=str(invoice_no)
        b1 = "MonthlyReport : " + str(file_name2)
        b="Date : "+str(d1)+"     "+str(hr)+":"+str(mn)
        file_direct=str(direct1)+str(file_name2)+".xlsx"
        file_name = str(directory)+str(bill_number)+"--"+str(cu_name)+".xlsx"
        c="Invoice No "+"0"+str(invoice_no)
        sheet = wb.active
        sheet.sheet_properties.fitToWidth = "False"
        sheet.page_setup.fitToHeight = 1
        #sheet.page_setup.fitToWidth = 0
        sheet.page_margins.top = 0
        sheet.page_margins.right =0
        sheet.page_margins.left =0
        sheet.page_margins.bottom = 1
        sheet.merge_cells('J2:L2')
        top_left_cell = sheet['J2']
        top_left_cell.value = str(c)
        top_left_cell.font = fontStyle
        top_left_cell.alignment = ali
        sheet.merge_cells('J3:N3')
        top_left_cell = sheet['J3']
        top_left_cell.value = str(b)
        top_left_cell.font = fontStyle
        sheet.merge_cells('J4:M4')
        top_left_cell = sheet['J4']
        top_left_cell.value = 'Liscense No:   20B/492/61/2019'
        top_left_cell.font = fontStyle
        sheet.merge_cells('J5:M5')
        top_left_cell = sheet['J5']
        top_left_cell.value = 'GUNA (M.P)'
        top_left_cell.font = fontStyle
        top_left_cell.alignment = ali
        sheet.merge_cells('A1:M1')
        top_left_cell = sheet['A1']
        top_left_cell.value = 'TAX INVOICE  -(CASH / CREDIT)'
        top_left_cell.font = fontStyle
        top_left_cell.alignment = ali
        sheet.merge_cells('D2:I2')
        top_left_cell = sheet['D2']
        top_left_cell.value = 'TO,'
        top_left_cell.font = fontStyle
        sheet.merge_cells('D3:I3')
        top_left_cell = sheet['D3']
        cust_name=cu_name
        top_left_cell.value = str(cust_name)
        top_left_cell.font = fontStyle
        sheet.merge_cells('D4:I4')
        top_left_cell = sheet['D4']
        cust_add=cu_add
        top_left_cell.value = str(cust_add)
        top_left_cell.font = fontStyle
        sheet.merge_cells('D5:I5')
        top_left_cell = sheet['D5']
        cust_mob="MOB NO  "+cu_pno
        top_left_cell.value = str(cust_mob)
        top_left_cell.font = fontStyle
        sheet.merge_cells('D6:I6')
        top_left_cell = sheet['D6']
        cust_gst="GST NO.  "+cu_gst
        top_left_cell.value = str(cust_gst)
        top_left_cell.font = fontStyle
        sheet.merge_cells('A2:B2')
        top_left_cell = sheet['A2']
        top_left_cell.value = 'FROM,'
        top_left_cell.font = fontStyle
        sheet.merge_cells('A3:C3')
        top_left_cell = sheet['A3']
        top_left_cell.value = "     SHRINATH AGENCY -GUNA "
        top_left_cell.font = fontStyle
        sheet.merge_cells('A4:C4')
        top_left_cell = sheet['A4']
        top_left_cell.value = "     KOTESWAR MANDIR GALI "
        top_left_cell.font = fontStyle
        top_left_cell.font = fontStyle
        sheet.merge_cells('A5:C5')
        top_left_cell = sheet['A5']
        top_left_cell.value = '     MOB NO. 9993370326 '
        top_left_cell.font = fontStyle
        sheet.merge_cells('A6:C6')
        top_left_cell = sheet['A6']
        top_left_cell.value = '     GST NO.  23AQEPM5243K1ZU'
        top_left_cell.font = fontStyle
        sheet.column_dimensions['A'].width =5
        sheet.column_dimensions['B'].width =7
        sheet.column_dimensions['C'].width = 23
        sheet.column_dimensions['D'].width = 7
        sheet.column_dimensions['E'].width = 6
        sheet.column_dimensions['F'].width = 7
        sheet.column_dimensions['G'].width = 6
        sheet.column_dimensions['H'].width = 7
        sheet.column_dimensions['I'].width = 6
        sheet.column_dimensions['J'].width = 6
        sheet.column_dimensions['K'].width = 6
        sheet.column_dimensions['L'].width = 6
        sheet.column_dimensions['M'].width = 7
        sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=13)
        sheet.cell(row=7, column=1).value="-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
        sheet.cell(row=7, column=1).font = fontStyle
        sheet['A8']='S.NO'
        sheet['B8']='HSN CODE'
        sheet['C8']='Product'
        sheet['D8']='MRP'
        sheet['E8']='QTTY'
        sheet['F8']='RATE'
        sheet['G8']='Dscnt%'
        sheet['H8']='TAX AMT'
        sheet['I8']='CGST%'
        sheet['J8']='CGST'
        sheet['K8']='SGST%'
        sheet['L8']='SGST'
        sheet['M8']='AMOUNT'
        for  j in range(1,14):
            sheet.cell(row=8, column=j).font = fontStyle
            sheet.cell(row=8, column=j).alignment = ali
        sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=13)
        sheet.cell(row=9, column=1).value="-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
        sheet.cell(row=9, column=1).font = fontStyle
        r = 1
        i = 0
        o=6
        l = 0
        i=0
        while l < len(products_list):
            sheet.cell(row=l+10, column=1).value = str(i+1)
            sheet.cell(row=l+10, column=2).value = str(product_code[i])
            sheet.cell(row=l+10, column=3).value = str(products_list[i])
            sheet.cell(row=l+10, column=4).value = str(product_mrp[i])
            sheet.cell(row=l+10, column=5).value = str(product_quantity[i])
            sheet.cell(row=l+10, column=6).value = str(product_rate[i])
            sheet.cell(row=l+10, column=7).value = str(product_dis[i])
            sheet.cell(row=l+10, column=8).value = str(product_price[i])
            sheet.cell(row=l+10, column=9).value = str(product_cgst[i])
            sheet.cell(row=l+10, column=10).value = str(product_cgstp[i])
            sheet.cell(row=l+10, column=11).value = str(product_sgst[i])
            sheet.cell(row=l+10, column=12).value = str(product_sgstp[i])
            sheet.cell(row=l+10, column=13).value = str(product_amount[i])
            sheet.cell(row=l+10, column=1).font = fontStyle
            sheet.cell(row=l+10, column=1).font = fontStyle
            sheet.cell(row=l+10, column=2).font = fontStyle
            sheet.cell(row=l+10, column=3).font = fontStyle
            sheet.cell(row=l+10, column=4).font = fontStyle
            sheet.cell(row=l+10, column=5).font = fontStyle
            sheet.cell(row=l+10, column=6).font = fontStyle
            sheet.cell(row=l+10, column=7).font = fontStyle
            sheet.cell(row=l+10, column=8).font = fontStyle
            sheet.cell(row=l+10, column=9).font = fontStyle
            sheet.cell(row=l+10, column=10).font = fontStyle
            sheet.cell(row=l+10, column=11).font = fontStyle
            sheet.cell(row=l+10, column=12).font = fontStyle
            sheet.cell(row=l+10, column=13).font = fontStyle
            sheet.cell(row=l+10, column=1).alignment = ali
            sheet.cell(row=l+10, column=2).alignment = ali
            sheet.cell(row=l+10, column=3).alignment = ali
            sheet.cell(row=l+10, column=4).alignment = ali
            sheet.cell(row=l+10, column=5).alignment = ali
            sheet.cell(row=l+10, column=6).alignment = ali
            sheet.cell(row=l+10, column=7).alignment = ali
            sheet.cell(row=l+10, column=8).alignment = ali
            sheet.cell(row=l+10, column=9).alignment = ali
            sheet.cell(row=l+10, column=10).alignment = ali
            sheet.cell(row=l+10, column=11).alignment = ali
            sheet.cell(row=l+10, column=12).alignment = ali
            sheet.cell(row=l+10, column=13).alignment = ali
            l=l+1
            i=i+1
        sheet.merge_cells(start_row=l+10, start_column=1, end_row=l+10, end_column=13)
        sheet.cell(row=l+10, column=1).value="------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
        sheet.cell(row=l+10, column=1).font = fontStyle
        sheet.cell(row=l+11, column=4).value = "TOTAL"
        sheet.cell(row=l+11, column=4).font = fontStyle
        amtw=0
        for row in sheet.iter_rows(min_row=10, min_col=13, max_row=l+10-1, max_col=13):
            for cell in row:
                amtw=amtw+float(cell.value)
        sheet.cell(row=l+11, column=13).value = str("%.2f"%amtw)
        sheet.cell(row=l+11, column=13).font = fontStyle
        sheet.cell(row=l+11, column=13).alignment = ali
        qtyw=0
        for row in sheet.iter_rows(min_row=10, min_col=5, max_row=l+9, max_col=5):
            for cell in row:
                qtyw=qtyw+float(cell.value)
        sheet.cell(row=l+11, column=5).value = str("%.2f"%qtyw)
        sheet.cell(row=l+11, column=5).font = fontStyle
        sheet.cell(row=l+11, column=5).alignment = ali
        tyw=0
        for row in sheet.iter_rows(min_row=10, min_col=8, max_row=l+9, max_col=8):
            for cell in row:
                tyw=tyw+float(cell.value)
        #sheet.cell(row=l+11, column=8).value = str("%.2f"%tyw)
        sheet.cell(row=l+11, column=8).font = fontStyle
        sheet.cell(row=l+11, column=8).alignment = ali
        yw=0
        for row in sheet.iter_rows(min_row=10, min_col=10, max_row=l+9, max_col=10):
            for cell in row:
                yw=yw+float(cell.value)
        sheet.cell(row=l+11, column=10).value = str("%.2f"%yw)
        sheet.cell(row=l+11, column=10).font = fontStyle
        sheet.cell(row=l+11, column=10).alignment = ali
        w=0
        for row in sheet.iter_rows(min_row=10, min_col=12, max_row=l+9, max_col=12):
            for cell in row:
                w=w+float(cell.value)
        sheet.cell(row=l+11, column=12).value = str("%.2f"%w)
        sheet.cell(row=l+11, column=12).alignment = ali
        sheet.cell(row=l+11, column=12).font = fontStyle
        sheet.merge_cells(start_row=l+12, start_column=1, end_row=12+l, end_column=3)
        sheet.cell(row=l+12, column=1).value="TAX DETAILS"
        sheet.cell(row=l+12, column=1).alignment = ali
        sheet.cell(row=l+12, column=1).font = fontStyle
        sheet.merge_cells(start_row=l+13, start_column=1, end_row=13+l, end_column=3)
        sheet.cell(row=l+13, column=1).value="--------------------------------------------------"
        sheet.cell(row=l+13, column=1).font = fontStyle
        sheet.merge_cells(start_row=l+14, start_column=1, end_row=l+14, end_column=3)
        sheet.cell(row=l+14, column=1).value="Tax    Tax%        Taxable         TaxAmt"
        sheet.cell(row=l+14, column=1).font = fontStyle
        sheet.cell(row=l+14, column=1).alignment = ali
        sheet.cell(row=l+15, column=1).value = "CGST"
        sheet.cell(row=l+15, column=1).font = fontStyle
        sheet.cell(row=l+15, column=1).alignment=ali
        sheet.cell(row=l+16, column=1).value = "SGST"
        sheet.cell(row=l+16, column=1).font = fontStyle
        sheet.cell(row=l+16, column=1).alignment=ali
        sheet.cell(row=l+15, column=2).value = "0%"
        sheet.cell(row=l+15, column=1).font = fontStyle
        sheet.cell(row=l+15, column=2).alignment=ali
        sheet.cell(row=l+16, column=2).value = "0%"
        sheet.cell(row=l+16, column=1).font = fontStyle
        sheet.cell(row=l+16, column=2).alignment=ali
        txc0=0
        txr0=0
        di=0
        for di in range(0,len(product_cgst)):
            if product_cgst[di] == 0:
                txc0=txc0+product_price[di]
                txr0 =product_cgstp[di]+txr0
            di=di+1
        sheet.cell(row=l+15, column=3).value = str("%.2f"%txc0)+"            " +str("%.2f"%txr0)
        sheet.cell(row=l+15, column=3).font = fontStyle
        sheet.cell(row=l+15, column=3).alignment=ali
        sheet.cell(row=l+16, column=3).value = str("%.2f"%txc0)+"            "+str("%.2f"%txr0)
        sheet.cell(row=l+16, column=3).font = fontStyle
        sheet.cell(row=l+16, column=3).alignment=ali
        sheet.cell(row=l+17, column=1).value = "CGST"
        sheet.cell(row=l+17, column=1).font = fontStyle
        sheet.cell(row=l+17, column=1).alignment=ali
        sheet.cell(row=l+18, column=1).value = "SGST"
        sheet.cell(row=l+18, column=1).font = fontStyle
        sheet.cell(row=l+18, column=1).alignment=ali
        sheet.cell(row=l+17, column=2).value = "6%"
        sheet.cell(row=l+17, column=2).font = fontStyle
        sheet.cell(row=l+17, column=2).alignment=ali
        sheet.cell(row=l+18, column=2).value = "6%"
        sheet.cell(row=l+18, column=2).font = fontStyle
        sheet.cell(row=l+18, column=2).alignment=ali
        sheet.cell(row=l+19, column=1).value = "CGST"
        sheet.cell(row=l+19, column=1).font = fontStyle
        sheet.cell(row=l+19, column=1).alignment=ali
        sheet.cell(row=l+20, column=1).value = "SGST"
        sheet.cell(row=l+20, column=1).font = fontStyle
        sheet.cell(row=l+20, column=1).alignment=ali
        sheet.cell(row=l+19, column=2).value = "9%"
        sheet.cell(row=l+19, column=2).font = fontStyle
        sheet.cell(row=l+19, column=2).alignment=ali
        sheet.cell(row=l+20, column=2).value = "9%"
        sheet.cell(row=l+20, column=2).font = fontStyle
        sheet.cell(row=l+20, column=2).alignment=ali
        txc9=0
        txr9=0
        di=0
        for di in range(0,len(product_cgst)):
            if product_cgst[di] == 9:
                txc9=txc9+product_price[di]
                txr9 =product_cgstp[di]+txr9
            di=di+1
        sheet.cell(row=l+19, column=3).value = str("%.2f"%txc9)+"            " +str("%.2f"%txr9)
        sheet.cell(row=l+19, column=3).font = fontStyle
        sheet.cell(row=l+19, column=3).alignment=ali
        sheet.cell(row=l+20, column=3).value = str("%.2f"%txc9)+"            "+str("%.2f"%txr9)
        sheet.cell(row=l+20, column=3).font = fontStyle
        sheet.cell(row=l+20, column=3).alignment=ali
        txr6=0
        txc6=0
        for di in range(0,len(product_cgst)):
            if product_cgst[di] == 6:
                txc6=txc6+product_price[di]
                txr6 =product_cgstp[di]+txr6
            di=di+1
        sheet.cell(row=l+17, column=3).value = str("%.2f"%txc6) +"          "+str("%.2f"%txr6)
        sheet.cell(row=l+17, column=3).font = fontStyle
        sheet.cell(row=l+17, column=3).alignment=ali
        sheet.cell(row=l+18, column=3).value = str("%.2f"%txc6) +"          "+str("%.2f"%txr6)
        sheet.cell(row=l+18, column=3).font = fontStyle
        sheet.cell(row=l+18, column=3).alignment=ali
        sheet.merge_cells(start_row=l+14, start_column=10, end_row=l+14, end_column=12)
        sheet.cell(row=l+14, column=10).value="Gross Amt"

        sheet.cell(row=l+14, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+14, start_column=13, end_row=l+14, end_column=13)
        sheet.cell(row=l+14, column=13).value=str("%.2f"%tyiw)
        sheet.cell(row=l+14, column=13).alignment = ali
        sheet.cell(row=l+14, column=13).font = fontStyle
        k=[]
        for i in range(0,len(product_dis)):
            k.append((float(product_dis[i])*product_rate[i]*product_quantity[i])/100.0)
        discountvalue=float("%.2f"%sum(k))
        sheet.merge_cells(start_row=l+15, start_column=10, end_row=l+15, end_column=12)
        sheet.cell(row=l+15, column=10).value="Scheme Disc Amt(-)"
        sheet.cell(row=l+15, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+15, start_column=13, end_row=l+15, end_column=13)
        sheet.cell(row=l+15, column=13).value=discountvalue
        sheet.cell(row=l+15, column=13).alignment = ali
        sheet.cell(row=l+15, column=13).font = fontStyle
        taxamts=(sum(product_cgstp)+sum(product_sgstp))
        sheet.merge_cells(start_row=l+16, start_column=10, end_row=l+16, end_column=12)
        sheet.cell(row=l+16, column=10).value="Tax Amt(+)"
        sheet.cell(row=l+16, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+16, start_column=13, end_row=l+16, end_column=13)
        sheet.cell(row=l+16, column=13).value=taxamts
        sheet.cell(row=l+16, column=13).alignment = ali
        sheet.cell(row=l+16, column=13).font = fontStyle
        totals=tyiw + taxamts - discountvalue
        jss=round(totals)
        js=str(jss)+".00"
        lml=(jss-totals)
        lnl=str(round(lml,2))
        sheet.merge_cells(start_row=l+17, start_column=10, end_row=l+17, end_column=12)
        sheet.cell(row=l+17, column=10).value="Round Off"
        sheet.cell(row=l+17, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+17, start_column=13, end_row=l+17, end_column=13)
        sheet.cell(row=l+17, column=13).value=lnl
        sheet.cell(row=l+17, column=13).alignment = ali
        sheet.cell(row=l+17, column=13).font = fontStyle
        sheet.merge_cells(start_row=l+18, start_column=10, end_row=l+18, end_column=12)
        sheet.cell(row=l+18, column=10).value="Net Payable"
        sheet.cell(row=l+18, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+18, start_column=13, end_row=l+18, end_column=14)
        sheet.cell(row=l+18, column=13).value=js
        sheet.cell(row=l+18, column=13).alignment = alia
        sheet.cell(row=l+18, column=13).font = fontamount
        sheet.merge_cells(start_row=l+19, start_column=10, end_row=l+19, end_column=13)
        sheet.cell(row=l+19, column=10).value="From Shrinath Agency Guna (M.P)"
        sheet.cell(row=l+19, column=10).alignment = ali
        sheet.cell(row=l+19, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+21, start_column=1, end_row=l+21, end_column=13)
        sheet.cell(row=l+21, column=1).value="DECLARATION: We here by certify that our rc under the GST ACT is in force on the date on which the sale of goods specified in this tax invoice."
        sheet.cell(row=l+21, column=1).alignment = alia
        sheet.cell(row=l+21, column=1).font = fontStyle

        #sheet.merge_cells(start_row=l+24, start_column=1, end_row=l+24, end_column=13)
        #sheet.cell(row=l+24, column=1).value="-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
        sheet.merge_cells(start_row=l+22, start_column=1, end_row=l+22, end_column=13)
        sheet.cell(row=l+22, column=1).value="टर्म्स एंड कंडीशन : "
        sheet.cell(row=l+22, column=1).alignment = alia
        sheet.cell(row=l+22, column=1).font = fontStyle
        sheet.merge_cells(start_row=l+23, start_column=1, end_row=l+23, end_column=13)
        sheet.cell(row=l+23, column=1).value="  1.डिलीवरी लेते समय माल चेक कर लेवे तथा किसी भी प्रकार के शार्ट माल की सूचना उसी दिन देवें अन्यथा हमारी कोई जबाबदारी  नही रहेगी।  "
        sheet.cell(row=l+23, column=1).alignment = alia
        sheet.cell(row=l+23, column=1).font = fontStyle
        sheet.merge_cells(start_row=l + 24, start_column=1, end_row=l + 24, end_column=13)
        sheet.cell(row=l + 24, column=1).value = "  2.पेमेन्ट 15  दिन में  भुगतान ना होने पर  2 % मासिक ब्याज  लिया जायेगा।"
        sheet.cell(row=l + 24, column=1).alignment = alia
        sheet.cell(row=l + 24, column=1).font = fontStyle
        sheet.merge_cells(start_row=l + 25, start_column=1, end_row=l + 25, end_column=13)
        sheet.cell(row=l + 25, column=1).value = "  3.न्यायक्षेत्र गुना रहेगा।  "
        sheet.cell(row=l + 25, column=1).alignment = alia
        sheet.cell(row=l + 25, column=1).font = fontStyle
        sheet.merge_cells(start_row=l + 26, start_column=1, end_row=l + 26, end_column=13)
        sheet.cell(row=l + 26, column=1).value = "  4.भूल चूक लेनी देनी। "
        sheet.cell(row=l + 26, column=1).alignment = alia
        sheet.cell(row=l + 26, column=1).font = fontStyle


        sheet.merge_cells(start_row=l+27, start_column=1, end_row=l+27, end_column=13)
        sheet.cell(row=l+27, column=1).value="        ----A Web Billing  Software Solution by Scyil Pvt Ltd Guna 9009323236----"
        sheet.cell(row=l+27, column=1).alignment = ali
        sheet.cell(row=l+27, column=1).font = fontScyil
        #wb.print_area = 'A1:C2'
        invoiced = Invoice(id=invoice_no+1,partyname=cu_name)
        invoiced.save()
        wb.save(file_name)
        #print(txr0, txc0, txr6, txc6, txr9, txc9)
        if (os.path.isfile(direct1 + file_name2 + '.xlsx')):
            xlsx = openpyxl.load_workbook(direct1 + file_name2 + '.xlsx')
            sheet1 = xlsx.active
            dimensions = sheet1.dimensions
            p = str(dimensions)
            po=int(p[p.find(":")+2:])
            l = po+1

            print(l)
            sno=2
            if(txc0>0):
                sheet1.cell(row=l, column=1).value = str(l-sno)
                sheet1.cell(row=l, column=2).value = str(cu_gst)
                sheet1.cell(row=l, column=3).value = str(cu_name)
                sheet1.cell(row=l, column=4).value = str(d1)
                sheet1.cell(row=l, column=5).value = str(invoice_no)
                sheet1.cell(row=l, column=6).value = str("%.2f" % tyiw)
                sheet1.cell(row=l, column=7).value = str("%.2f" % txc0)
                sheet1.cell(row=l, column=8).value = str(0.00)
                l = l + 1
            if (txc6 > 0):
                sheet1.cell(row=l, column=1).value = str(l-sno)
                sheet1.cell(row=l, column=2).value = str(cu_gst)
                sheet1.cell(row=l, column=3).value = str(cu_name)
                sheet1.cell(row=l, column=4).value = str(d1)
                sheet1.cell(row=l, column=5).value = str(invoice_no)
                sheet1.cell(row=l, column=6).value = str("%.2f" % tyiw)
                sheet1.cell(row=l, column=7).value = str("%.2f" % txc6)
                sheet1.cell(row=l, column=8).value = str(12.00)
                l=l+1
            if (txc9>0):
                sheet1.cell(row=l, column=1).value = str(l - sno)
                sheet1.cell(row=l, column=2).value = str(cu_gst)
                sheet1.cell(row=l, column=3).value = str(cu_name)
                sheet1.cell(row=l, column=4).value = str(d1)
                sheet1.cell(row=l, column=5).value = str(invoice_no)
                sheet1.cell(row=l, column=6).value = str("%.2f" % tyiw)
                sheet1.cell(row=l, column=7).value = str("%.2f" % txc9)
                sheet1.cell(row=l, column=8).value = str(18.00)
                l=l+1
            xlsx.save(file_direct)
        else:
            wb1= Workbook()
            sheet1 = wb1.active
            if not os.path.exists(direct1):
                os.makedirs(direct1)
            sheet1.sheet_properties.fitToWidth = "FALSE"
            sheet1.page_setup.fitToHeight = 1
            sheet1.page_setup.fitToWidth = 1
            sheet1.page_margins.left = 0
            sheet1.page_margins.right = 0
            sheet1.column_dimensions['A'].width =5
            sheet1.column_dimensions['B'].width =20
            sheet1.column_dimensions['C'].width = 20
            sheet1.column_dimensions['D'].width = 10
            sheet1.column_dimensions['E'].width = 10
            sheet1.column_dimensions['F'].width = 10
            sheet1.column_dimensions['G'].width = 10
            sheet1.column_dimensions['H'].width = 10
            sheet1['B1'] = 'SHRINATH AGENCY'
            temp="Month  "+file_name2
            sheet1['C1'] = str(temp)
            sheet1['A2'] = 'S NO.'
            sheet1['B2'] = 'GST NO'
            sheet1['C2'] = 'Particulars'
            sheet1['D2'] = 'Date'
            sheet1['E2'] = 'Invoice NO.'
            sheet1['F2'] = 'Gross Total'
            sheet1['G2'] = 'Taxable Amt'
            sheet1['H2'] = 'Rate'
            ind=0
            l=0
            if (txc0 > 0):
                sheet1.cell(row=l+3, column=1).value = str(ind+1)
                sheet1.cell(row=l+3, column=2).value = str(cu_gst)
                sheet1.cell(row=l+3, column=3).value = str(cu_name)
                sheet1.cell(row=l+3, column=4).value = str(d1)
                sheet1.cell(row=l+3, column=5).value = str(invoice_no)
                sheet1.cell(row=l+3, column=6).value = str("%.2f" % tyiw)
                sheet1.cell(row=l+3, column=7).value = str("%.2f" % txc0)
                sheet1.cell(row=l+3, column=8).value = str("0.00")
                l = l + 1
                ind=ind+1
            if (txc6 > 0):
                sheet1.cell(row=l+3, column=1).value = str(ind+1)
                sheet1.cell(row=l+3, column=2).value = str(cu_gst)
                sheet1.cell(row=l+3, column=3).value = str(cu_name)
                sheet1.cell(row=l+3, column=4).value = str(d1)
                sheet1.cell(row=l+3, column=5).value = str(invoice_no)
                sheet1.cell(row=l+3, column=6).value = str("%.2f" % tyiw)
                sheet1.cell(row=l+3, column=7).value = str("%.2f" % txc6)
                sheet1.cell(row=l+3, column=8).value = str("12.00")
                l = l + 1
                ind=ind+1
            if (txc9 > 0):
                sheet1.cell(row=l+3, column=1).value = str(ind+1)
                sheet1.cell(row=l+3, column=2).value = str(cu_gst)
                sheet1.cell(row=l+3, column=3).value = str(cu_name)
                sheet1.cell(row=l+3, column=4).value = str(d1)
                sheet1.cell(row=l+3, column=5).value = str(invoice_no)
                sheet1.cell(row=l+3, column=6).value = str("%.2f" % tyiw)
                sheet1.cell(row=l+3, column=7).value = str("%.2f" % txc9)
                sheet1.cell(row=l+3, column=8).value = str("18.00")
                l = l + 1
                ind=ind+1
            wb1.save(file_direct)
        if(flag_print==1):
            os.startfile(file_name, "print")
        return render(request,"print.html", {'curl': curl})
    else:
        return render(request,"print.html", {'curl': curl})


